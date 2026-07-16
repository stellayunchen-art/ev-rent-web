# export_benchmarks.py
# ============================================================
# 作用：把 Excel 知识库导出成 benchmarks.csv，供 Streamlit 网页版使用
# 运行环境：MacBook（本地）
# 运行时机：每次有新站点过会后，重新跑一次，然后把 benchmarks.csv 提交到 GitHub
# ============================================================

import openpyxl
import csv
from pathlib import Path

# ─── 修改这里 ──────────────────────────────────────────────
_BASE     = Path("/Users/stellachan/Documents/JV/租金管理表/租金评估")
_UPDATED  = _BASE / "AI租金自动评估" / "换电站选址租金知识库_租金评估结果AI自动化更新_updated.xlsx"
_ORIGINAL = _BASE / "换电站知识库" / "换电站选址租金知识库_租金评估结果AI自动化更新.xlsx"
INPUT_FILE = _UPDATED if _UPDATED.exists() else _ORIGINAL
OUTPUT_CSV = Path(__file__).parent / "benchmarks.csv"   # 输出到 coze网页版/ 目录
SHEET_NAME = "案例知识库"
# ───────────────────────────────────────────────────────────

DATA_START = 2

FIELDS = ["name", "city", "district", "coord", "bc_type", "area_type", "road_cond", "road_type", "unit_rent", "bound_rent", "audit_date"]

# ⚠️ 不再用硬编码列号！2026-07-15发现Excel删掉"场地类型"列后，O列以后整体左移一列，
# 硬编码列号（C_UNIT=15等）读到了错误的列（unit_rent读成了边界、bound_rent读成了站点名称）。
# 改为按表头名称查列——只要表头文字不变，列增删移动都不受影响。
# 表头名→字段名映射（Excel第1行的列标题）：
HEADER_MAP = {
    "站点名称":       "name",
    "内审日期":       "audit_date",
    "城市":           "city",
    "坐标":           "coord",
    "商圈类型":       "bc_type",
    "区域类型":       "area_type",
    "道路条件":       "road_cond",
    "道路类型":       "road_type",
    "单车位租金":     "unit_rent",
    "单车位租金边界": "bound_rent",
}
# 行政区列的表头历史上有两种叫法，任一命中即可
DISTRICT_HEADERS = ("行政区", "区/县街道/镇", "区/县")


def safe(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("#N/A", "None", "nan") else s


def resolve_columns(ws) -> dict:
    """读第1行表头，返回 {字段名: 列号(1-based)}。找不到关键列时报错退出。"""
    header_to_col = {}
    for c in range(1, ws.max_column + 1):
        h = safe(ws.cell(1, c).value)
        if h:
            header_to_col[h] = c
    cols = {}
    for header, field in HEADER_MAP.items():
        if header not in header_to_col:
            raise KeyError(f"Excel表头缺少「{header}」列，当前表头：{list(header_to_col)}")
        cols[field] = header_to_col[header]
    # 行政区：兼容多种表头名
    for dh in DISTRICT_HEADERS:
        if dh in header_to_col:
            cols["district"] = header_to_col[dh]
            break
    else:
        cols["district"] = None  # 找不到就留空，不阻断导出
    return cols


def main():
    if not INPUT_FILE.exists():
        print(f"❌ 找不到文件：{INPUT_FILE}")
        print("请确认 Excel 路径是否正确")
        return

    print(f"加载 Excel：{INPUT_FILE.name}")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ 找不到 Sheet「{SHEET_NAME}」，当前 Sheet 列表：{wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    cols = resolve_columns(ws)
    print(f"按表头解析到列号：{cols}")

    def cell(r, field):
        c = cols.get(field)
        return ws.cell(r, c).value if c else None

    rows = []
    skipped = 0
    for r in range(DATA_START, ws.max_row + 1):
        name  = safe(cell(r, "name"))
        coord = safe(cell(r, "coord"))

        # 跳过：无名称 / 坐标为空 / 待评估（新增未过会站点不作为对标）
        if not name or "," not in coord:
            skipped += 1
            continue

        rows.append({
            "name":      name,
            "city":      safe(cell(r, "city")),
            "district":  safe(cell(r, "district")),
            "coord":     coord,
            "bc_type":   safe(cell(r, "bc_type")),
            "area_type": safe(cell(r, "area_type")),
            "road_cond": safe(cell(r, "road_cond")),
            "road_type": safe(cell(r, "road_type")),
            "unit_rent": safe(cell(r, "unit_rent")),
            "bound_rent":safe(cell(r, "bound_rent")),
            "audit_date":safe(cell(r, "audit_date"))[:10],  # 内审日期，仅保留年月日
        })

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    print(f"✅ 已导出 {len(rows)} 条记录 → {OUTPUT_CSV}")
    print(f"   （跳过 {skipped} 行空行/无坐标行）")
    print()
    print("下一步：")
    print("  1. git add benchmarks.csv")
    print("  2. git commit -m 'update benchmarks'")
    print("  3. git push")
    print("  → Streamlit Cloud 会自动拉取更新，网页版即时生效")


if __name__ == "__main__":
    main()
